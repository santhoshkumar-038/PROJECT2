#In this project we are decreasing the price of the things by using python with openpyxl
#as updating manually it takes so much of time,
#we are accessing through the excel by this manuall and decreasing each price by 90% by using formula price*0.9

import openpyxl as xl      #importing openpyxl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx') #In this step we are loading our excel
                                           #by using xl.load_workbook('...xlsx') workbook means the excel sheet we
                                           #are trying to read or do any operations

sheet = wb['Sheet1'] #we are accessing our excel sheet by using wb['Sheet1'] we should use caps S as it is a case
                     #sensitivew

#now we are learning how to access to cell in sheet as cell:it is combination of rows and column
cell = sheet['a1']
#another way to access by using cell function

cell = sheet.cell(1,1)  #(1,1) is row and column
#print(cell.value) #it prints the value of cell

#print(sheet.max_row) #it prints the no of rows we used in excel sheet...

for row in range(2, sheet.max_row+1):  #this step we are starting with 2 row as 1 row
                                       #consists of transaction id and other values so i am ignoring it
     
     cell = sheet.cell(row, 3)         #in this sheet.cell(row, 3) and i am iterating from 2 row and column 3
                                       #it gives values of our dollars 
     #print(cell.value)
     corrected_value = cell.value*0.9  #it multiplies each value with 0.9 as we are decreasing by 90%

     #we are creating another row and column to represent another row
     corrected_value_cell = sheet.cell(row,4)
     corrected_value_cell.value = corrected_value

#let's add chart to the sheet so we can do this by importing BarChart, Reference

values = Reference(sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col =4) #we are taking from 2 to sheet.max_row which is 4

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('transactions2.xlsx')




















